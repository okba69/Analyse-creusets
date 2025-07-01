import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from io import BytesIO
from collections import Counter

st.set_page_config(page_title="Analyse Creusets", layout="wide")
st.title("🔍 Analyse et détection de sets/anomalies")
st.markdown("Dépose ton fichier Excel, puis clique sur **Analyser**. Les seuils sont préconfigurés.")

CLEAN_THRESHOLD   = 60
SET_THRESHOLD     = 80
ANOMALY_THRESHOLD = 70
MIN_DROP_COUNT    = 30  # <-- Ajoute ici le seuil de repassage sous 70 pour déclencher un nouveau set

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    data_cols = df.columns[1:57]
    df[data_cols] = df[data_cols].applymap(
        lambda x: ""
        if pd.notna(x) and isinstance(x, (int, float)) and (x in [99, 100] or x < CLEAN_THRESHOLD)
        else x
    )
    for i in range(len(df)):
        if (df.loc[i, data_cols] == "").sum() >= 40:
            df.loc[i, data_cols] = ""
    for i in range(len(df)-1):
        cur = pd.to_numeric(df.loc[i, data_cols], errors='coerce')
        nxt = pd.to_numeric(df.loc[i+1, data_cols], errors='coerce')
        if ((cur - nxt) >= 15).sum() >= 15:
            df.loc[i, data_cols] = ""
    for col in data_cols:
        for i in range(1, len(df)-1):
            if df.at[i-1, col] == "" and df.at[i+1, col] == "":
                df.at[i, col] = ""
    return df

def detect_sets_and_anomalies(df: pd.DataFrame):
    data_cols = list(df.columns[1:57])
    set_starts = []
    set_count = 0
    in_set = False
    anomalies_by_set = {}
    anomaly_cells = []
    meta = []
    current = set()
    last = 0
    dropped_flags = {ci: False for ci in range(len(data_cols))}
    dropped_since_last_set = set()  # <-- pour compter les emplacements descendus sous 70

    for idx in range(len(df)):
        vals = pd.to_numeric(df.loc[idx, data_cols], errors='coerce')
        cnt80 = (vals > SET_THRESHOLD).sum()
        cnt70 = (vals > ANOMALY_THRESHOLD).sum()

        # Met à jour dropped_flags pour chaque colonne
        for ci, col in enumerate(data_cols):
            v = pd.to_numeric(df.at[idx, col], errors='coerce')
            # repère chaque descente SOUS 70
            if pd.notna(v) and v < ANOMALY_THRESHOLD:
                dropped_flags[ci] = True
                dropped_since_last_set.add(ci)  # Ajoute cet emplacement à la liste "descendus sous 70"

        if not in_set:
            # SEULEMENT si 30 emplacements sont repassés sous 70, on autorise la détection du nouveau set
            if len(dropped_since_last_set) >= MIN_DROP_COUNT and cnt80 >= 40:
                set_count += 1
                try:
                    ts = pd.to_datetime(df.iloc[idx,0], errors='coerce')
                    if pd.notnull(ts):
                        date = ts.strftime("%d/%m/%Y %H:%M")
                    else:
                        date = "Inconnu"
                except:
                    date = "Inconnu"
                # sauvegarde set précédent
                if last > 0 and current:
                    anomalies_by_set[last] = sorted(current)
                meta.append({"Set": set_count, "Date": date})
                last = set_count
                current = set()
                in_set = True
                dropped_flags = {ci: False for ci in range(len(data_cols))}
                dropped_since_last_set = set()  # On RESET le compteur après démarrage du set
            else:
                for ci, col in enumerate(data_cols):
                    if not dropped_flags[ci]:
                        continue
                    v = pd.to_numeric(df.at[idx, col], errors='coerce')
                    if pd.notna(v) and v >= SET_THRESHOLD:
                        if idx+1 < len(df):
                            nv = pd.to_numeric(df.at[idx+1, col], errors='coerce')
                            if pd.notna(nv) and nv < SET_THRESHOLD:
                                continue
                        colnum = ci + 1
                        if colnum not in current:
                            current.add(colnum)
                            anomaly_cells.append((idx+2, ci+2))
        else:
            if cnt70 < ANOMALY_THRESHOLD:
                in_set = False

    if last > 0 and current:
        anomalies_by_set[last] = sorted(current)
    return set_starts, anomaly_cells, meta, anomalies_by_set

def to_excel(df, set_starts, anomaly_cells, meta, anomalies_by_set, ranking):
    wb = Workbook()
    ws = wb.active
    ws.title = "Données nettoyées"

    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY HH:MM')
    if "date_style" not in wb.named_styles:
        wb.add_named_style(date_style)
    dates_col = pd.to_datetime(df.iloc[:,0], errors='coerce')
    for r in range(len(df)):
        cell = ws.cell(row=r+1, column=1)
        v = dates_col.iloc[r]
        if pd.notnull(v):
            cell.value = v
            cell.style = date_style
        else:
            cell.value = None
            cell.style = date_style
        for c in range(1, len(df.columns)):
            ws.cell(row=r+1, column=c+1, value=df.iloc[r, c])
    orange = PatternFill("solid", fgColor="FFA500")
    blue   = PatternFill("solid", fgColor="ADD8E6")
    for row_idx in set_starts:
        for cell in ws[row_idx]:
            cell.fill = orange
    for row, col in anomaly_cells:
        ws.cell(row=row, column=col).fill = blue
    ws2 = wb.create_sheet("Résumé")
    ws2.append(["Set", "Date", "Nb anomalies"])
    total_anomalies = 0
    for m in meta:
        s = m["Set"]
        nb = len(anomalies_by_set.get(s, []))
        total_anomalies += nb
        ws2.append([s, m["Date"], nb])
    ws2.append([])
    ws2.append(["Total anomalies", total_anomalies])
    ws2.append([])
    ws2.append(["Classement complet Emplacements","Occurrences"])
    for val, cnt in ranking:
        ws2.append([val, cnt])
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(i)].width = 5.5
    for col_cells in ws2.columns:
        ws2.column_dimensions[get_column_letter(col_cells[0].column)].width = 15
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

uploaded = st.file_uploader("Téléverse ton fichier .xlsx", type=["xlsx"])
analyse = st.button("Analyser")

if uploaded and analyse:
    df = pd.read_excel(uploaded, engine='openpyxl')
    df_clean = clean_data(df.copy())
    set_starts, anomaly_cells, meta, anomalies_by_set = detect_sets_and_anomalies(df_clean)
    all_anoms = [x for vals in anomalies_by_set.values() for x in vals]
    ranking = Counter(all_anoms).most_common()
    recap = pd.DataFrame([
        {"Set": m["Set"], "Date": m["Date"], "Nb anomalies": len(anomalies_by_set.get(m["Set"], []))}
        for m in meta
    ]).set_index("Set")
    st.markdown("## Récapitulatif des sets")
    st.table(recap)
    total = recap["Nb anomalies"].sum()
    st.markdown(f"**Total anomalies sur tous les sets : {total}**")
    st.markdown("## Classement complet des emplacements changés le plus souvent")
    df_ranking = pd.DataFrame(ranking, columns=["Emplacement","Occurrences"]).set_index("Emplacement")
    st.table(df_ranking)
    excel_bytes = to_excel(df_clean, set_starts, anomaly_cells, meta, anomalies_by_set, ranking)
    st.download_button(
        "📥 Télécharger le rapport Excel",
        data=excel_bytes,
        file_name="analyse_creusets_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
